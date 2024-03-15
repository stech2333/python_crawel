from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from urllib.parse import quote
import time
import openpyxl

def read_keywords_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    #i = 0
    keywords = []

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        #i = i+1
        if row[0]:  # 确保单元格不为空
            keywords.append(row[0])
        #if i>15:    #这个是调试时候用的，一开始试了16个，好像没问题
        #    break
    #for key in keywords:
    #    print(key+"\n")
    return keywords

def get_company_href(driver,keyword):
    # 指定EdgeDriver路径
    try:
        # 打开天眼查搜索页面
        url = "https://www.tianyancha.com/search?key=" + quote(keyword)
        driver.get(url)
        
        # 等待页面加载完成
        #time.sleep(50)  # 根据你的网络速度调整等待时间
        # 使用显式等待等待元素出现
        wait = WebDriverWait(driver, 10)  # 等待最多10秒
        a_tag = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.index_alink__zcia5.link-click')))
        href = a_tag.get_attribute('href')
        return href  # 返回链接
    finally:
        # 关闭浏览器
        pass
def get_time(driver,url):
    driver.get(url)
    
    wait = WebDriverWait(driver, 10)  # 等待最多10秒
    div_tag = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'index_com-header-push-main__EyhAZ')))
        
    div_text = div_tag.text
    print(div_text+"\n")  # 打印div标签的文本内容
    return div_text



if __name__ == "__main__":

    try:
        #获取驱动
        driver_path = 'msedgedriver.exe'
        service = Service(executable_path=driver_path)
        driver = webdriver.Edge(service=service)

        #新建文件
        wb = Workbook()
        ws = wb.active

        #自己登陆
        driver.get("https://www.tianyancha.com/login")
        print("自己登陆进去，有60s时间登录")
        time.sleep(60)  # 等待登录
        keywords = read_keywords_from_excel("企业名录.xlsx")
        #hrefs = []
        #results = []
        for keyword in keywords: 
            if get_company_href(driver,keyword):  # 确保href不是None
#                hrefs.append(href)
#        for a in hrefs:
                out_time = get_time(driver,get_company_href(driver,keyword))
                #results.append((keyword, out_time))
                ws.append([keyword, out_time])
                wb.save("result.xlsx")

        #for keyword, out_time in results:
              # 将每个(keyword, out_time)对添加到新行

        # 保存工作簿到文件
        
    finally:
        print("运行结束")